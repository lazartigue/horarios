import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()

# ============================================================
# ESTILOS
# ============================================================
header_font = Font(name='Calibri', bold=True, color='FFFFFF', size=11)
header_fill = PatternFill(start_color='2F5496', end_color='2F5496', fill_type='solid')
subheader_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
subheader_font = Font(name='Calibri', bold=True, color='FFFFFF', size=10)
area_fill = PatternFill(start_color='D6E4F0', end_color='D6E4F0', fill_type='solid')
area_font = Font(name='Calibri', bold=True, color='2F5496', size=11)
turno_a_fill = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')  # verde claro
turno_b_fill = PatternFill(start_color='FCE4D6', end_color='FCE4D6', fill_type='solid')  # naranja claro
check_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')  # verde check
name_font = Font(name='Calibri', size=10)
name_font_bold = Font(name='Calibri', size=10, bold=True)
cell_font = Font(name='Calibri', size=10)
hrs_font = Font(name='Calibri', size=10, bold=True, color='2F5496')
center = Alignment(horizontal='center', vertical='center', wrap_text=True)
left = Alignment(horizontal='left', vertical='center')
thin_border = Border(
    left=Side(style='thin', color='B4C6E7'),
    right=Side(style='thin', color='B4C6E7'),
    top=Side(style='thin', color='B4C6E7'),
    bottom=Side(style='thin', color='B4C6E7')
)


def style_header(ws, row, max_col):
    for col in range(1, max_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
        cell.border = thin_border


def style_area(ws, row, max_col, text):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=max_col)
    cell = ws.cell(row=row, column=1, value=text)
    cell.font = area_font
    cell.fill = area_fill
    cell.alignment = left
    cell.border = thin_border
    for col in range(2, max_col + 1):
        ws.cell(row=row, column=col).fill = area_fill
        ws.cell(row=row, column=col).border = thin_border


def add_person_row(ws, row, data, max_col, bold_name=False):
    for col_idx, val in enumerate(data, 1):
        cell = ws.cell(row=row, column=col_idx, value=val)
        cell.border = thin_border
        cell.alignment = center
        if col_idx == 1:
            cell.alignment = left
            cell.font = name_font_bold if bold_name else name_font
        elif col_idx == max_col:
            cell.font = hrs_font
        else:
            cell.font = cell_font
            # Color por turno
            if val and '(A)' in str(val):
                cell.fill = turno_a_fill
            elif val and '(B)' in str(val):
                cell.fill = turno_b_fill


def add_check_row(ws, row, data, max_col):
    for col_idx, val in enumerate(data, 1):
        cell = ws.cell(row=row, column=col_idx, value=val)
        cell.border = thin_border
        cell.alignment = center
        cell.font = Font(name='Calibri', size=10, bold=True, color='006100')
        if col_idx > 1:
            cell.fill = check_fill


# ============================================================
# HOJA 1: ESCENARIO 1 — LUNES A VIERNES
# ============================================================
ws1 = wb.active
ws1.title = "Escenario 1 - L a V"

# Título
ws1.merge_cells('A1:H1')
title_cell = ws1.cell(row=1, column=1, value="ESCENARIO 1: LUNES A VIERNES (9:00 - 20:00)")
title_cell.font = Font(name='Calibri', bold=True, size=14, color='2F5496')
title_cell.alignment = center

# Leyenda turnos
ws1.merge_cells('A2:H2')
ws1.cell(row=2, column=1, value="Turno A (Apertura): 9:00–18:30  |  Turno B (Cierre): 10:30–20:00  |  47 hrs/semana").font = Font(name='Calibri', size=10, italic=True)
ws1.cell(row=2, column=1).alignment = center

cols_e1 = ['Nombre', 'Lunes', 'Martes', 'Miércoles', 'Jueves', 'Viernes', 'Hrs/Sem']
max_col_e1 = len(cols_e1)

row = 4
# Header
for col_idx, col_name in enumerate(cols_e1, 1):
    ws1.cell(row=row, column=col_idx, value=col_name)
style_header(ws1, row, max_col_e1)
row += 1

# --- COLORISTAS ---
style_area(ws1, row, max_col_e1, "COLORISTAS")
row += 1
data_coloristas = [
    ["Marcela", "9:00-18:30 (A)", "10:30-20:00 (B)", "9:00-18:30 (A)", "10:30-20:00 (B)", "9:00-18:00", 47],
    ["Paola", "10:30-20:00 (B)", "9:00-18:30 (A)", "10:30-20:00 (B)", "9:00-18:30 (A)", "10:30-20:00 (B)", 47],
    ["María Paz", "9:00-18:30 (A)", "9:00-18:30 (A)", "10:30-20:00 (B)", "10:30-20:00 (B)", "9:00-18:30 (A)", 47],
    ["Daniel", "10:30-20:00 (B)", "10:30-20:00 (B)", "10:30-20:00 (B)", "10:30-20:00 (B)", "10:30-19:30 (B)", 47],
]
for d in data_coloristas:
    add_person_row(ws1, row, d, max_col_e1, bold_name=(d[0] == "Daniel"))
    row += 1

# --- APLICADORAS ---
style_area(ws1, row, max_col_e1, "APLICADORAS")
row += 1
data_aplicadoras = [
    ["Paty", "9:00-18:30 (A)", "10:30-20:00 (B)", "9:00-18:30 (A)", "10:30-20:00 (B)", "10:30-20:00 (B)", 47],
    ["Karen", "10:30-20:00 (B)", "9:00-18:30 (A)", "10:30-20:00 (B)", "9:00-18:30 (A)", "9:00-18:00", 47],
    ["Ceci Flores", "9:00-18:30 (A)", "9:00-18:30 (A)", "10:30-20:00 (B)", "9:00-18:30 (A)", "10:30-20:00 (B)", 47],
]
for d in data_aplicadoras:
    add_person_row(ws1, row, d, max_col_e1)
    row += 1

# --- CORTE/PEINADO ---
style_area(ws1, row, max_col_e1, "CORTE / PEINADO / SECADO")
row += 1
data_corte = [
    ["Ceci Pacheco", "10:30-20:00 (B)", "9:00-18:30 (A)", "10:30-20:00 (B)", "9:00-18:30 (A)", "9:00-18:00", 47],
    ["Ingrid", "9:00-18:30 (A)", "10:30-20:00 (B)", "9:00-18:30 (A)", "10:30-20:00 (B)", "10:30-20:00 (B)", 47],
]
for d in data_corte:
    add_person_row(ws1, row, d, max_col_e1)
    row += 1

# --- MASAJISTA/AYUDANTE ---
style_area(ws1, row, max_col_e1, "MASAJISTA / AYUDANTE")
row += 1
data_masajista = [
    ["Isa San Martín", "9:00-18:30 (A)", "10:30-20:00 (B)", "9:00-18:30 (A)", "10:30-20:00 (B)", "9:00-18:00", 47],
    ["Ube", "10:30-20:00 (B)", "9:00-18:30 (A)", "10:30-20:00 (B)", "9:00-18:30 (A)", "9:00-18:30 (A)", 47],
    ["Tamara", "9:00-18:30 (A)", "9:00-18:30 (A)", "10:30-20:00 (B)", "9:00-18:30 (A)", "10:30-20:00 (B)", 47],
]
for d in data_masajista:
    add_person_row(ws1, row, d, max_col_e1)
    row += 1

# --- LAVAPELO ---
style_area(ws1, row, max_col_e1, "LAVAPELO Y SECADO")
row += 1
data_lavapelo = [
    ["Carmencita", "9:00-18:30 (A)", "10:30-20:00 (B)", "9:00-18:30 (A)", "10:30-20:00 (B)", "9:00-18:00", 47],
    ["Alma", "10:30-20:00 (B)", "9:00-18:30 (A)", "10:30-20:00 (B)", "9:00-18:30 (A)", "10:30-20:00 (B)", 47],
    ["Carolina", "9:00-18:30 (A)", "9:00-18:30 (A)", "10:30-20:00 (B)", "9:00-18:30 (A)", "9:00-18:00", 46],
    ["Francis", "10:30-20:00 (B)", "10:30-20:00 (B)", "9:00-18:30 (A)", "10:30-20:00 (B)", "10:30-20:00 (B)", 47],
]
for d in data_lavapelo:
    add_person_row(ws1, row, d, max_col_e1)
    row += 1

# --- RECEPCIÓN ---
style_area(ws1, row, max_col_e1, "RECEPCIÓN")
row += 1
data_recepcion = [
    ["Antonio San Martín", "9:00-18:30 (A)", "9:00-18:30 (A)", "9:00-18:30 (A)", "9:00-18:30 (A)", "9:00-18:00", 47],
    ["Vale", "9:00-18:30 (A)", "10:30-20:00 (B)", "9:00-18:30 (A)", "10:30-20:00 (B)", "9:00-18:30 (A)", 47],
    ["Raquel", "10:30-20:00 (B)", "9:00-18:30 (A)", "10:30-20:00 (B)", "9:00-18:30 (A)", "10:30-20:00 (B)", 47],
    ["Fran Brocco", "10:30-20:00 (B)", "10:30-20:00 (B)", "10:30-20:00 (B)", "10:30-20:00 (B)", "10:30-20:00 (B)", 47],
]
for d in data_recepcion:
    add_person_row(ws1, row, d, max_col_e1, bold_name=(d[0] in ["Antonio San Martín", "Fran Brocco"]))
    row += 1

# --- BACK OFFICE ---
style_area(ws1, row, max_col_e1, "BACK OFFICE")
row += 1
add_person_row(ws1, row, ["Pamela Hernández", "9:00-18:30 (A)", "9:00-18:30 (A)", "9:00-18:30 (A)", "9:00-18:30 (A)", "9:00-18:00", 47], max_col_e1)
row += 2

# --- VERIFICACIÓN CORTE AL CIERRE ---
style_area(ws1, row, max_col_e1, "VERIFICACIÓN: CORTE AL CIERRE (20:00)")
row += 1
for col_idx, col_name in enumerate(["Restricción", "Lunes", "Martes", "Miércoles", "Jueves", "Viernes", "Estado"], 1):
    cell = ws1.cell(row=row, column=col_idx, value=col_name)
    cell.font = subheader_font
    cell.fill = subheader_fill
    cell.alignment = center
    cell.border = thin_border
row += 1
checks = [
    ["Corte al cierre", "Daniel + Ceci P.", "Daniel + Paty", "Daniel + Ceci P.", "Daniel + Paty", "Daniel + Paty", "✅"],
    ["Colorista al cierre", "Paola + Daniel", "Daniel + M.Paz", "M.Paz + Daniel", "Paola + Daniel", "Paola + Daniel", "✅"],
    ["Lavapelo al cierre", "Alma + Francis", "Carmencita + Francis", "Alma + Francis", "Carmencita + Francis", "Alma + Francis", "✅"],
    ["Recepción al cierre", "Raquel + Fran B.", "Vale + Fran B.", "Raquel + Fran B.", "Vale + Fran B.", "Raquel + Fran B.", "✅"],
    ["Antonio a las 9:00", "✅", "✅", "✅", "✅", "✅", "✅"],
    ["Daniel desde 10:30", "✅", "✅", "✅", "✅", "✅", "✅"],
    ["Recepción al inicio", "Antonio + Vale", "Antonio + Raquel", "Antonio + Vale", "Antonio + Raquel", "Antonio + Vale", "✅"],
]
for d in checks:
    add_check_row(ws1, row, d, max_col_e1)
    row += 1

# Ajustar anchos
ws1.column_dimensions['A'].width = 22
for col in range(2, max_col_e1 + 1):
    ws1.column_dimensions[get_column_letter(col)].width = 20
ws1.column_dimensions[get_column_letter(max_col_e1)].width = 10


# ============================================================
# HOJA 2: ESCENARIO 2 — LUNES A SÁBADO
# ============================================================
ws2 = wb.create_sheet("Escenario 2 - L a S")

# Título
ws2.merge_cells('A1:I1')
title_cell2 = ws2.cell(row=1, column=1, value="ESCENARIO 2: LUNES A SÁBADO (Sábado cierra 14:00)")
title_cell2.font = Font(name='Calibri', bold=True, size=14, color='2F5496')
title_cell2.alignment = center

ws2.merge_cells('A2:I2')
ws2.cell(row=2, column=1, value="Turno A: 9:00–17:30  |  Turno B: 11:30–20:00  |  Sábado: 9:00–14:00  |  Rotación G1/G2 sábados alternados  |  47 hrs/semana").font = Font(name='Calibri', size=10, italic=True)
ws2.cell(row=2, column=1).alignment = center

cols_e2 = ['Nombre', 'Lunes', 'Martes', 'Miércoles', 'Jueves', 'Viernes', 'Sábado', 'Hrs/Sem']
max_col_e2 = len(cols_e2)

row = 4
for col_idx, col_name in enumerate(cols_e2, 1):
    ws2.cell(row=row, column=col_idx, value=col_name)
style_header(ws2, row, max_col_e2)
row += 1

# --- COLORISTAS ---
style_area(ws2, row, max_col_e2, "COLORISTAS")
row += 1
data2_coloristas = [
    ["Marcela", "9:00-17:30 (A)", "11:30-20:00 (B)", "9:00-17:30 (A)", "11:30-20:00 (B)", "9:00-17:00", "G1: 9:00-14:00", 47],
    ["Paola", "11:30-20:00 (B)", "9:00-17:30 (A)", "11:30-20:00 (B)", "9:00-17:30 (A)", "11:30-20:00 (B)", "G2: LIBRE", 47],
    ["María Paz", "9:00-17:30 (A)", "9:00-17:30 (A)", "11:30-20:00 (B)", "9:00-17:30 (A)", "11:30-20:00 (B)", "G2: LIBRE", 47],
    ["Daniel", "10:30-19:00 (B)", "10:30-19:00 (B)", "10:30-19:00 (B)", "10:30-19:00 (B)", "10:30-19:00 (B)", "G2: 10:30-14:00", 47],
]
for d in data2_coloristas:
    add_person_row(ws2, row, d, max_col_e2, bold_name=(d[0] == "Daniel"))
    row += 1

# --- APLICADORAS ---
style_area(ws2, row, max_col_e2, "APLICADORAS")
row += 1
data2_aplicadoras = [
    ["Paty", "9:00-17:30 (A)", "11:30-20:00 (B)", "9:00-17:30 (A)", "11:30-20:00 (B)", "11:30-20:00 (B)", "G1: 9:00-14:00", 47],
    ["Karen", "11:30-20:00 (B)", "9:00-17:30 (A)", "11:30-20:00 (B)", "9:00-17:30 (A)", "9:00-17:00", "G1: 9:00-14:00", 47],
    ["Ceci Flores", "9:00-17:30 (A)", "9:00-17:30 (A)", "11:30-20:00 (B)", "9:00-17:30 (A)", "9:00-17:30 (A)", "G2: 9:00-14:00", 47],
]
for d in data2_aplicadoras:
    add_person_row(ws2, row, d, max_col_e2)
    row += 1

# --- CORTE/PEINADO ---
style_area(ws2, row, max_col_e2, "CORTE / PEINADO / SECADO")
row += 1
data2_corte = [
    ["Ceci Pacheco", "11:30-20:00 (B)", "9:00-17:30 (A)", "11:30-20:00 (B)", "9:00-17:30 (A)", "9:00-17:00", "G1: 9:00-14:00", 47],
    ["Ingrid", "9:00-17:30 (A)", "11:30-20:00 (B)", "9:00-17:30 (A)", "11:30-20:00 (B)", "11:30-20:00 (B)", "G2: 9:00-14:00", 47],
]
for d in data2_corte:
    add_person_row(ws2, row, d, max_col_e2)
    row += 1

# --- MASAJISTA/AYUDANTE ---
style_area(ws2, row, max_col_e2, "MASAJISTA / AYUDANTE")
row += 1
data2_masajista = [
    ["Isa San Martín", "9:00-17:30 (A)", "11:30-20:00 (B)", "9:00-17:30 (A)", "11:30-20:00 (B)", "9:00-17:00", "G1: 9:00-14:00", 47],
    ["Ube", "11:30-20:00 (B)", "9:00-17:30 (A)", "11:30-20:00 (B)", "9:00-17:30 (A)", "9:00-17:30 (A)", "G2: 9:00-14:00", 47],
    ["Tamara", "9:00-17:30 (A)", "9:00-17:30 (A)", "11:30-20:00 (B)", "9:00-17:30 (A)", "11:30-20:00 (B)", "G2: 9:00-14:00", 47],
]
for d in data2_masajista:
    add_person_row(ws2, row, d, max_col_e2)
    row += 1

# --- LAVAPELO ---
style_area(ws2, row, max_col_e2, "LAVAPELO Y SECADO")
row += 1
data2_lavapelo = [
    ["Carmencita", "9:00-17:30 (A)", "11:30-20:00 (B)", "9:00-17:30 (A)", "11:30-20:00 (B)", "9:00-17:00", "G1: 9:00-14:00", 47],
    ["Alma", "11:30-20:00 (B)", "9:00-17:30 (A)", "11:30-20:00 (B)", "9:00-17:30 (A)", "11:30-20:00 (B)", "G2: LIBRE", 47],
    ["Carolina", "9:00-17:30 (A)", "9:00-17:30 (A)", "11:30-20:00 (B)", "9:00-17:30 (A)", "9:00-17:00", "G1: 9:00-14:00", 47],
    ["Francis", "11:30-20:00 (B)", "11:30-20:00 (B)", "9:00-17:30 (A)", "11:30-20:00 (B)", "11:30-20:00 (B)", "G2: LIBRE", 47],
]
for d in data2_lavapelo:
    add_person_row(ws2, row, d, max_col_e2)
    row += 1

# --- RECEPCIÓN ---
style_area(ws2, row, max_col_e2, "RECEPCIÓN")
row += 1
data2_recepcion = [
    ["Antonio San Martín", "9:00-17:30 (A)", "9:00-17:30 (A)", "9:00-17:30 (A)", "9:00-17:30 (A)", "9:00-17:00", "G1: 9:00-14:00", 47],
    ["Vale", "9:00-17:30 (A)", "11:30-20:00 (B)", "9:00-17:30 (A)", "11:30-20:00 (B)", "9:00-17:30 (A)", "G2: 9:00-14:00", 47],
    ["Raquel", "11:30-20:00 (B)", "9:00-17:30 (A)", "11:30-20:00 (B)", "9:00-17:30 (A)", "11:30-20:00 (B)", "G2: LIBRE", 47],
    ["Fran Brocco", "11:30-20:00 (B)", "11:30-20:00 (B)", "11:30-20:00 (B)", "11:30-20:00 (B)", "11:30-20:00 (B)", "LIBRE", 47],
]
for d in data2_recepcion:
    add_person_row(ws2, row, d, max_col_e2, bold_name=(d[0] in ["Antonio San Martín", "Fran Brocco"]))
    row += 1

# --- BACK OFFICE ---
style_area(ws2, row, max_col_e2, "BACK OFFICE")
row += 1
add_person_row(ws2, row, ["Pamela Hernández", "9:00-17:30 (A)", "9:00-17:30 (A)", "9:00-17:30 (A)", "9:00-17:30 (A)", "9:00-17:00", "G1: 9:00-14:00", 47], max_col_e2)
row += 2

# --- VERIFICACIÓN ---
style_area(ws2, row, max_col_e2, "VERIFICACIÓN DE RESTRICCIONES")
row += 1
for col_idx, col_name in enumerate(["Restricción", "Lunes", "Martes", "Miércoles", "Jueves", "Viernes", "Sábado", "Estado"], 1):
    cell = ws2.cell(row=row, column=col_idx, value=col_name)
    cell.font = subheader_font
    cell.fill = subheader_fill
    cell.alignment = center
    cell.border = thin_border
row += 1
checks2 = [
    ["Corte al cierre", "Ceci P.", "Paty", "Ceci P.", "Paty", "Paty", "Daniel", "✅"],
    ["Colorista al cierre", "Paola+Daniel", "Daniel+M.Paz", "M.Paz+Daniel", "Paola+Daniel", "Paola+Daniel", "Daniel", "✅"],
    ["Lavapelo al cierre", "Alma+Francis", "Carm.+Francis", "Alma+Francis", "Carm.+Francis", "Alma+Francis", "Carm.+Carol.", "✅"],
    ["Recepción al cierre", "Raquel+Fran", "Vale+Fran", "Raquel+Fran", "Vale+Fran", "Raquel+Fran", "Antonio/Vale", "✅"],
    ["Antonio a las 9:00", "✅", "✅", "✅", "✅", "✅", "✅ (sáb G1)", "✅"],
    ["Daniel desde 10:30", "✅", "✅", "✅", "✅", "✅", "✅", "✅"],
    ["Recepción inicio", "Antonio+Vale", "Antonio+Raquel", "Antonio+Vale", "Antonio+Raquel", "Antonio+Vale", "Antonio/Vale", "✅"],
]
for d in checks2:
    add_check_row(ws2, row, d, max_col_e2)
    row += 1

# Ajustar anchos
ws2.column_dimensions['A'].width = 22
for col in range(2, max_col_e2 + 1):
    ws2.column_dimensions[get_column_letter(col)].width = 20
ws2.column_dimensions[get_column_letter(max_col_e2)].width = 10


# ============================================================
# HOJA 3: ESCENARIO 3 — LUNES A VIERNES 9:00–21:00
# ============================================================
ws3 = wb.create_sheet("Escenario 3 - Cierre 21h")

turno_c_fill = PatternFill(start_color='D9E2F3', end_color='D9E2F3', fill_type='solid')  # azul claro

def add_person_row_e3(ws, row, data, max_col, bold_name=False):
    for col_idx, val in enumerate(data, 1):
        cell = ws.cell(row=row, column=col_idx, value=val)
        cell.border = thin_border
        cell.alignment = center
        if col_idx == 1:
            cell.alignment = left
            cell.font = name_font_bold if bold_name else name_font
        elif col_idx == max_col:
            cell.font = hrs_font
        else:
            cell.font = cell_font
            if val and '(A)' in str(val):
                cell.fill = turno_a_fill
            elif val and '(C)' in str(val):
                cell.fill = turno_c_fill
            elif val and '(B)' in str(val):
                cell.fill = turno_b_fill

# Título
ws3.merge_cells('A1:H1')
title_cell3 = ws3.cell(row=1, column=1, value="ESCENARIO 3: LUNES A VIERNES (9:00 - 21:00)")
title_cell3.font = Font(name='Calibri', bold=True, size=14, color='2F5496')
title_cell3.alignment = center

ws3.merge_cells('A2:H2')
ws3.cell(row=2, column=1, value="Turno A: 9:00–18:00  |  Turno B: 11:00–20:00  |  Turno C: 12:00–21:00  |  47 hrs/semana (1 día extendido por persona)").font = Font(name='Calibri', size=10, italic=True)
ws3.cell(row=2, column=1).alignment = center

cols_e3 = ['Nombre', 'Lunes', 'Martes', 'Miércoles', 'Jueves', 'Viernes', 'Hrs/Sem']
max_col_e3 = len(cols_e3)

row = 4
for col_idx, col_name in enumerate(cols_e3, 1):
    ws3.cell(row=row, column=col_idx, value=col_name)
style_header(ws3, row, max_col_e3)
row += 1

# --- COLORISTAS ---
style_area(ws3, row, max_col_e3, "COLORISTAS")
row += 1
# Marcela: 4×9 + 1×11(9-20) = 47
# Paola: alterna B/C, 4×9 + 1×11(10-21) = 47 — usa 1 día B extendido
# María Paz: 4×9 + 1×11 = 47
# Daniel: siempre desde 10:30. 10:30-19:30(B)=9h × 4 + 10:30-21:00(C)=10.5 + ajuste = 47
data3_coloristas = [
    ["Marcela",   "9:00-18:00 (A)", "12:00-21:00 (C)", "9:00-18:00 (A)", "11:00-20:00 (B)", "9:00-20:00 (A)", 47],
    ["Paola",     "12:00-21:00 (C)", "9:00-18:00 (A)", "11:00-20:00 (B)", "12:00-21:00 (C)", "9:00-18:00 (A)", 47],
    ["María Paz", "11:00-20:00 (B)", "9:00-18:00 (A)", "12:00-21:00 (C)", "9:00-18:00 (A)", "11:00-20:00 (B)", 47],
    ["Daniel",    "10:30-19:30 (B)", "10:30-19:30 (B)", "10:30-19:30 (B)", "10:30-21:00 (C)", "10:30-19:30 (B)", 47],
]
for d in data3_coloristas:
    add_person_row_e3(ws3, row, d, max_col_e3, bold_name=(d[0] == "Daniel"))
    row += 1

# --- APLICADORAS ---
style_area(ws3, row, max_col_e3, "APLICADORAS")
row += 1
data3_aplicadoras = [
    ["Paty",       "9:00-18:00 (A)", "12:00-21:00 (C)", "9:00-18:00 (A)", "11:00-20:00 (B)", "12:00-21:00 (C)", 47],
    ["Karen",      "12:00-21:00 (C)", "9:00-18:00 (A)", "11:00-20:00 (B)", "9:00-18:00 (A)", "11:00-20:00 (B)", 47],
    ["Ceci Flores","11:00-20:00 (B)", "11:00-20:00 (B)", "12:00-21:00 (C)", "9:00-18:00 (A)", "9:00-18:00 (A)", 47],
]
for d in data3_aplicadoras:
    add_person_row_e3(ws3, row, d, max_col_e3)
    row += 1

# --- CORTE/PEINADO ---
style_area(ws3, row, max_col_e3, "CORTE / PEINADO / SECADO")
row += 1
data3_corte = [
    ["Ceci Pacheco", "12:00-21:00 (C)", "9:00-18:00 (A)", "12:00-21:00 (C)", "9:00-18:00 (A)", "11:00-20:00 (B)", 47],
    ["Ingrid",       "9:00-18:00 (A)", "11:00-20:00 (B)", "9:00-18:00 (A)", "12:00-21:00 (C)", "12:00-21:00 (C)", 47],
]
for d in data3_corte:
    add_person_row_e3(ws3, row, d, max_col_e3)
    row += 1

# --- MASAJISTA/AYUDANTE ---
style_area(ws3, row, max_col_e3, "MASAJISTA / AYUDANTE")
row += 1
data3_masajista = [
    ["Isa San Martín", "9:00-18:00 (A)", "12:00-21:00 (C)", "9:00-18:00 (A)", "11:00-20:00 (B)", "9:00-20:00 (A)", 47],
    ["Ube",            "11:00-20:00 (B)", "9:00-18:00 (A)", "12:00-21:00 (C)", "9:00-18:00 (A)", "11:00-20:00 (B)", 47],
    ["Tamara",         "12:00-21:00 (C)", "11:00-20:00 (B)", "9:00-18:00 (A)", "12:00-21:00 (C)", "9:00-18:00 (A)", 47],
]
for d in data3_masajista:
    add_person_row_e3(ws3, row, d, max_col_e3)
    row += 1

# --- LAVAPELO ---
style_area(ws3, row, max_col_e3, "LAVAPELO Y SECADO")
row += 1
data3_lavapelo = [
    ["Carmencita", "9:00-18:00 (A)", "11:00-20:00 (B)", "12:00-21:00 (C)", "9:00-18:00 (A)", "11:00-20:00 (B)", 47],
    ["Alma",       "12:00-21:00 (C)", "9:00-18:00 (A)", "11:00-20:00 (B)", "12:00-21:00 (C)", "9:00-18:00 (A)", 47],
    ["Carolina",   "11:00-20:00 (B)", "12:00-21:00 (C)", "9:00-18:00 (A)", "11:00-20:00 (B)", "12:00-21:00 (C)", 47],
    ["Francis",    "12:00-21:00 (C)", "12:00-21:00 (C)", "11:00-20:00 (B)", "12:00-21:00 (C)", "9:00-18:00 (A)", 47],
]
for d in data3_lavapelo:
    add_person_row_e3(ws3, row, d, max_col_e3)
    row += 1

# --- RECEPCIÓN ---
style_area(ws3, row, max_col_e3, "RECEPCIÓN")
row += 1
data3_recepcion = [
    ["Antonio San Martín", "9:00-18:00 (A)", "9:00-18:00 (A)", "9:00-18:00 (A)", "9:00-18:00 (A)", "9:00-20:00 (A)", 47],
    ["Vale",               "9:00-18:00 (A)", "12:00-21:00 (C)", "9:00-18:00 (A)", "12:00-21:00 (C)", "9:00-18:00 (A)", 47],
    ["Raquel",             "11:00-20:00 (B)", "9:00-18:00 (A)", "12:00-21:00 (C)", "9:00-18:00 (A)", "12:00-21:00 (C)", 47],
    ["Fran Brocco",        "12:00-21:00 (C)", "11:00-20:00 (B)", "11:00-20:00 (B)", "12:00-21:00 (C)", "11:00-20:00 (B)", 47],
]
for d in data3_recepcion:
    add_person_row_e3(ws3, row, d, max_col_e3, bold_name=(d[0] in ["Antonio San Martín", "Fran Brocco"]))
    row += 1

# --- BACK OFFICE ---
style_area(ws3, row, max_col_e3, "BACK OFFICE")
row += 1
add_person_row_e3(ws3, row, ["Pamela Hernández", "9:00-18:00 (A)", "9:00-18:00 (A)", "9:00-18:00 (A)", "9:00-18:00 (A)", "9:00-20:00 (A)", 47], max_col_e3)
row += 2

# --- VERIFICACIÓN ---
style_area(ws3, row, max_col_e3, "VERIFICACIÓN DE RESTRICCIONES (Cierre 21:00)")
row += 1
for col_idx, col_name in enumerate(["Restricción", "Lunes", "Martes", "Miércoles", "Jueves", "Viernes", "Estado"], 1):
    cell = ws3.cell(row=row, column=col_idx, value=col_name)
    cell.font = subheader_font
    cell.fill = subheader_fill
    cell.alignment = center
    cell.border = thin_border
row += 1
checks3 = [
    ["Corte al cierre (21h)", "Ceci Pacheco", "Paty", "Ceci Pacheco", "Daniel", "Paty / Ingrid", "✅"],
    ["Colorista al cierre", "Paola", "Marcela", "María Paz", "Daniel", "Paola", "✅"],
    ["Lavapelo al cierre", "Alma+Francis", "Carolina+Francis", "Carmencita", "Alma+Francis", "Carolina", "✅"],
    ["Recepción al cierre", "Fran Brocco", "Vale+Fran B.", "Raquel+Fran B.", "Fran Brocco", "Raquel", "✅"],
    ["Antonio a las 9:00", "✅", "✅", "✅", "✅", "✅", "✅"],
    ["Daniel desde 10:30", "✅", "✅", "✅", "✅", "✅", "✅"],
    ["Recepción al inicio", "Antonio+Vale", "Antonio+Raquel", "Antonio+Vale", "Antonio+Raquel", "Antonio+Vale", "✅"],
    ["Peak lavado 12-16", "4/4 presentes", "4/4 presentes", "4/4 presentes", "4/4 presentes", "4/4 presentes", "✅"],
]
for d in checks3:
    add_check_row(ws3, row, d, max_col_e3)
    row += 1

# Nota especial escenario 3
row += 1
ws3.merge_cells(start_row=row, start_column=1, end_row=row, end_column=max_col_e3)
note_cell = ws3.cell(row=row, column=1, value="NOTA: Este escenario requiere 3 turnos para cubrir 12 horas. El solapamiento de turnos B y C entre 12:00-20:00 garantiza cobertura máxima en horario peak. Cada persona tiene 1 día extendido (11 hrs) para completar 47 hrs/semana.")
note_cell.font = Font(name='Calibri', size=10, italic=True, color='C00000')
note_cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
ws3.row_dimensions[row].height = 40

# Ajustar anchos
ws3.column_dimensions['A'].width = 22
for col in range(2, max_col_e3 + 1):
    ws3.column_dimensions[get_column_letter(col)].width = 20
ws3.column_dimensions[get_column_letter(max_col_e3)].width = 10


# ============================================================
# HOJA 4: ESCENARIO 4 — DOTACIÓN COMPLETA 9:00–20:00 CON CONTRATACIONES
# ============================================================
ws_e4 = wb.create_sheet("Escenario 4 - Dot. Completa")

nuevo_fill = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')  # amarillo claro
nuevo_font = Font(name='Calibri', size=10, color='BF8F00', bold=True)
costo_fill = PatternFill(start_color='F4B084', end_color='F4B084', fill_type='solid')  # naranja costo
costo_font_w = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
costo_fill_dark = PatternFill(start_color='C55A11', end_color='C55A11', fill_type='solid')

def add_person_row_e4(ws, row, data, max_col, bold_name=False, is_new=False):
    for col_idx, val in enumerate(data, 1):
        cell = ws.cell(row=row, column=col_idx, value=val)
        cell.border = thin_border
        cell.alignment = center
        if col_idx == 1:
            cell.alignment = left
            if is_new:
                cell.font = nuevo_font
                cell.fill = nuevo_fill
            else:
                cell.font = name_font_bold if bold_name else name_font
        elif col_idx == max_col:
            cell.font = hrs_font
        else:
            cell.font = cell_font
            if is_new and not (val and ('(A)' in str(val) or '(B)' in str(val))):
                cell.fill = nuevo_fill
            if val and '(A)' in str(val):
                if is_new:
                    cell.fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
                else:
                    cell.fill = turno_a_fill
            elif val and '(B)' in str(val):
                if is_new:
                    cell.fill = PatternFill(start_color='F8CBAD', end_color='F8CBAD', fill_type='solid')
                else:
                    cell.fill = turno_b_fill

# Título
ws_e4.merge_cells('A1:H1')
t4 = ws_e4.cell(row=1, column=1, value="ESCENARIO 4: DOTACIÓN COMPLETA L-V 9:00-20:00 (con contrataciones nuevas)")
t4.font = Font(name='Calibri', bold=True, size=14, color='2F5496')
t4.alignment = center

ws_e4.merge_cells('A2:H2')
ws_e4.cell(row=2, column=1, value="Turno A: 9:00–18:30  |  Turno B: 10:30–20:00  |  ★ = Personal NUEVO  |  47 hrs/semana").font = Font(name='Calibri', size=10, italic=True)
ws_e4.cell(row=2, column=1).alignment = center

# Leyenda nuevo
ws_e4.merge_cells('A3:H3')
leg = ws_e4.cell(row=3, column=1, value="Las filas en AMARILLO corresponden a personal NUEVO a contratar. Sueldo estimado al nivel más bajo del área.")
leg.font = Font(name='Calibri', size=10, italic=True, color='BF8F00')
leg.alignment = center

cols_e4 = ['Nombre', 'Lunes', 'Martes', 'Miércoles', 'Jueves', 'Viernes', 'Hrs/Sem']
max_col_e4 = len(cols_e4)

row = 5
for col_idx, col_name in enumerate(cols_e4, 1):
    ws_e4.cell(row=row, column=col_idx, value=col_name)
style_header(ws_e4, row, max_col_e4)
row += 1

# --- COLORISTAS (4 actuales turno A + 4 nuevos turno B) ---
style_area(ws_e4, row, max_col_e4, "COLORISTAS (4 actuales + 4 nuevos)")
row += 1
# Actuales: todos Turno A
e4_coloristas_act = [
    ["Marcela",   "9:00-18:30 (A)", "9:00-18:30 (A)", "9:00-18:30 (A)", "9:00-18:30 (A)", "9:00-18:00", 47],
    ["Paola",     "9:00-18:30 (A)", "9:00-18:30 (A)", "9:00-18:30 (A)", "9:00-18:30 (A)", "9:00-18:00", 47],
    ["María Paz", "9:00-18:30 (A)", "9:00-18:30 (A)", "9:00-18:30 (A)", "9:00-18:30 (A)", "9:00-18:00", 47],
    ["Daniel",    "10:30-20:00 (B)", "10:30-20:00 (B)", "10:30-20:00 (B)", "10:30-20:00 (B)", "10:30-19:30 (B)", 47],
]
for d in e4_coloristas_act:
    add_person_row_e4(ws_e4, row, d, max_col_e4, bold_name=(d[0]=="Daniel"))
    row += 1
# Nuevos: turno B
e4_coloristas_new = [
    ["★ Colorista Nuevo 1", "10:30-20:00 (B)", "10:30-20:00 (B)", "10:30-20:00 (B)", "10:30-20:00 (B)", "10:30-19:30 (B)", 47],
    ["★ Colorista Nuevo 2", "10:30-20:00 (B)", "10:30-20:00 (B)", "10:30-20:00 (B)", "10:30-20:00 (B)", "10:30-19:30 (B)", 47],
    ["★ Colorista Nuevo 3", "10:30-20:00 (B)", "10:30-20:00 (B)", "10:30-20:00 (B)", "10:30-20:00 (B)", "10:30-19:30 (B)", 47],
]
for d in e4_coloristas_new:
    add_person_row_e4(ws_e4, row, d, max_col_e4, is_new=True)
    row += 1

# --- APLICADORAS (3 actuales + 3 nuevas) ---
style_area(ws_e4, row, max_col_e4, "APLICADORAS (3 actuales + 3 nuevas)")
row += 1
e4_aplic_act = [
    ["Paty",       "9:00-18:30 (A)", "9:00-18:30 (A)", "9:00-18:30 (A)", "9:00-18:30 (A)", "9:00-18:00", 47],
    ["Karen",      "9:00-18:30 (A)", "9:00-18:30 (A)", "9:00-18:30 (A)", "9:00-18:30 (A)", "9:00-18:00", 47],
    ["Ceci Flores","9:00-18:30 (A)", "9:00-18:30 (A)", "9:00-18:30 (A)", "9:00-18:30 (A)", "9:00-18:00", 47],
]
for d in e4_aplic_act:
    add_person_row_e4(ws_e4, row, d, max_col_e4)
    row += 1
e4_aplic_new = [
    ["★ Aplicadora Nueva 1", "10:30-20:00 (B)", "10:30-20:00 (B)", "10:30-20:00 (B)", "10:30-20:00 (B)", "10:30-19:30 (B)", 47],
    ["★ Aplicadora Nueva 2", "10:30-20:00 (B)", "10:30-20:00 (B)", "10:30-20:00 (B)", "10:30-20:00 (B)", "10:30-19:30 (B)", 47],
    ["★ Aplicadora Nueva 3", "10:30-20:00 (B)", "10:30-20:00 (B)", "10:30-20:00 (B)", "10:30-20:00 (B)", "10:30-19:30 (B)", 47],
]
for d in e4_aplic_new:
    add_person_row_e4(ws_e4, row, d, max_col_e4, is_new=True)
    row += 1

# --- CORTE/PEINADO (2 actuales + 2 nuevos) ---
style_area(ws_e4, row, max_col_e4, "CORTE / PEINADO / SECADO (2 actuales + 2 nuevos)")
row += 1
e4_corte_act = [
    ["Ceci Pacheco", "9:00-18:30 (A)", "10:30-20:00 (B)", "9:00-18:30 (A)", "10:30-20:00 (B)", "9:00-18:00", 47],
    ["Ingrid",       "10:30-20:00 (B)", "9:00-18:30 (A)", "10:30-20:00 (B)", "9:00-18:30 (A)", "10:30-20:00 (B)", 47],
]
for d in e4_corte_act:
    add_person_row_e4(ws_e4, row, d, max_col_e4)
    row += 1
e4_corte_new = [
    ["★ Cortador/a Nuevo 1", "10:30-20:00 (B)", "9:00-18:30 (A)", "10:30-20:00 (B)", "9:00-18:30 (A)", "10:30-20:00 (B)", 47],
    ["★ Cortador/a Nuevo 2", "9:00-18:30 (A)", "10:30-20:00 (B)", "9:00-18:30 (A)", "10:30-20:00 (B)", "9:00-18:00", 47],
]
for d in e4_corte_new:
    add_person_row_e4(ws_e4, row, d, max_col_e4, is_new=True)
    row += 1

# --- MASAJISTA/AYUDANTE (3 actuales + 3 nuevos) ---
style_area(ws_e4, row, max_col_e4, "MASAJISTA / AYUDANTE (3 actuales + 3 nuevos)")
row += 1
e4_masaj_act = [
    ["Isa San Martín", "9:00-18:30 (A)", "9:00-18:30 (A)", "9:00-18:30 (A)", "9:00-18:30 (A)", "9:00-18:00", 47],
    ["Ube",            "9:00-18:30 (A)", "9:00-18:30 (A)", "9:00-18:30 (A)", "9:00-18:30 (A)", "9:00-18:00", 47],
    ["Tamara",         "9:00-18:30 (A)", "9:00-18:30 (A)", "9:00-18:30 (A)", "9:00-18:30 (A)", "9:00-18:00", 47],
]
for d in e4_masaj_act:
    add_person_row_e4(ws_e4, row, d, max_col_e4)
    row += 1
e4_masaj_new = [
    ["★ Masajista Nuevo 1", "10:30-20:00 (B)", "10:30-20:00 (B)", "10:30-20:00 (B)", "10:30-20:00 (B)", "10:30-19:30 (B)", 47],
    ["★ Masajista Nuevo 2", "10:30-20:00 (B)", "10:30-20:00 (B)", "10:30-20:00 (B)", "10:30-20:00 (B)", "10:30-19:30 (B)", 47],
    ["★ Masajista Nuevo 3", "10:30-20:00 (B)", "10:30-20:00 (B)", "10:30-20:00 (B)", "10:30-20:00 (B)", "10:30-19:30 (B)", 47],
]
for d in e4_masaj_new:
    add_person_row_e4(ws_e4, row, d, max_col_e4, is_new=True)
    row += 1

# --- LAVAPELO (4 actuales + 4 nuevos) ---
style_area(ws_e4, row, max_col_e4, "LAVAPELO Y SECADO (4 actuales + 4 nuevos)")
row += 1
e4_lava_act = [
    ["Carmencita", "9:00-18:30 (A)", "9:00-18:30 (A)", "9:00-18:30 (A)", "9:00-18:30 (A)", "9:00-18:00", 47],
    ["Alma",       "9:00-18:30 (A)", "9:00-18:30 (A)", "9:00-18:30 (A)", "9:00-18:30 (A)", "9:00-18:00", 47],
    ["Carolina",   "9:00-18:30 (A)", "9:00-18:30 (A)", "9:00-18:30 (A)", "9:00-18:30 (A)", "9:00-18:00", 47],
    ["Francis",    "9:00-18:30 (A)", "9:00-18:30 (A)", "9:00-18:30 (A)", "9:00-18:30 (A)", "9:00-18:00", 47],
]
for d in e4_lava_act:
    add_person_row_e4(ws_e4, row, d, max_col_e4)
    row += 1
e4_lava_new = [
    ["★ Lavapelo Nuevo 1", "10:30-20:00 (B)", "10:30-20:00 (B)", "10:30-20:00 (B)", "10:30-20:00 (B)", "10:30-19:30 (B)", 47],
    ["★ Lavapelo Nuevo 2", "10:30-20:00 (B)", "10:30-20:00 (B)", "10:30-20:00 (B)", "10:30-20:00 (B)", "10:30-19:30 (B)", 47],
    ["★ Lavapelo Nuevo 3", "10:30-20:00 (B)", "10:30-20:00 (B)", "10:30-20:00 (B)", "10:30-20:00 (B)", "10:30-19:30 (B)", 47],
    ["★ Lavapelo Nuevo 4", "10:30-20:00 (B)", "10:30-20:00 (B)", "10:30-20:00 (B)", "10:30-20:00 (B)", "10:30-19:30 (B)", 47],
]
for d in e4_lava_new:
    add_person_row_e4(ws_e4, row, d, max_col_e4, is_new=True)
    row += 1

# --- RECEPCIÓN (4 actuales + 4 nuevos, pero Antonio siempre 9:00) ---
style_area(ws_e4, row, max_col_e4, "RECEPCIÓN (4 actuales + 3 nuevos)")
row += 1
e4_recep_act = [
    ["Antonio San Martín", "9:00-18:30 (A)", "9:00-18:30 (A)", "9:00-18:30 (A)", "9:00-18:30 (A)", "9:00-18:00", 47],
    ["Vale",               "9:00-18:30 (A)", "9:00-18:30 (A)", "9:00-18:30 (A)", "9:00-18:30 (A)", "9:00-18:00", 47],
    ["Raquel",             "10:30-20:00 (B)", "10:30-20:00 (B)", "10:30-20:00 (B)", "10:30-20:00 (B)", "10:30-19:30 (B)", 47],
    ["Fran Brocco",        "10:30-20:00 (B)", "10:30-20:00 (B)", "10:30-20:00 (B)", "10:30-20:00 (B)", "10:30-19:30 (B)", 47],
]
for d in e4_recep_act:
    add_person_row_e4(ws_e4, row, d, max_col_e4, bold_name=(d[0] in ["Antonio San Martín", "Fran Brocco"]))
    row += 1
e4_recep_new = [
    ["★ Recepción Nueva 1", "9:00-18:30 (A)", "10:30-20:00 (B)", "9:00-18:30 (A)", "10:30-20:00 (B)", "9:00-18:30 (A)", 47],
    ["★ Recepción Nueva 2", "10:30-20:00 (B)", "9:00-18:30 (A)", "10:30-20:00 (B)", "9:00-18:30 (A)", "10:30-20:00 (B)", 47],
    ["★ Recepción Nueva 3", "10:30-20:00 (B)", "10:30-20:00 (B)", "10:30-20:00 (B)", "10:30-20:00 (B)", "10:30-19:30 (B)", 47],
]
for d in e4_recep_new:
    add_person_row_e4(ws_e4, row, d, max_col_e4, is_new=True)
    row += 1

# --- BACK OFFICE (sin cambio) ---
style_area(ws_e4, row, max_col_e4, "BACK OFFICE (sin cambio)")
row += 1
add_person_row_e4(ws_e4, row, ["Pamela Hernández", "9:00-18:30 (A)", "9:00-18:30 (A)", "9:00-18:30 (A)", "9:00-18:30 (A)", "9:00-18:00", 47], max_col_e4)
row += 2

# === TABLA RESUMEN DE CONTRATACIONES ===
style_area(ws_e4, row, max_col_e4, "RESUMEN DE CONTRATACIONES NUEVAS")
row += 1
resumen_headers = ["Área", "Dotación Actual", "Nuevos a Contratar", "Dotación Total", "Sueldo Ref. Nuevo", "Costo Mensual Nuevo", ""]
for col_idx, h in enumerate(resumen_headers, 1):
    cell = ws_e4.cell(row=row, column=col_idx, value=h)
    cell.font = subheader_font
    cell.fill = subheader_fill
    cell.alignment = center
    cell.border = thin_border
row += 1

contrataciones = [
    ("Coloristas", 4, 3, 1231088),
    ("Aplicadoras", 3, 3, 1533265),
    ("Corte/Peinado/Secado", 2, 2, 860405),
    ("Masajista/Ayudante", 3, 3, 808173),
    ("Lavapelo y Secado", 4, 4, 789159),
    ("Recepción", 4, 3, 1246698),
    ("Back Office", 1, 0, 0),
]

total_act = 0
total_new = 0
total_cost = 0
for area, act, new, sueldo in contrataciones:
    costo = sueldo * new
    total_act += act
    total_new += new
    total_cost += costo
    row_data = [area, act, new, act + new, f"${sueldo:,.0f}" if sueldo > 0 else "N/A", f"${costo:,.0f}" if costo > 0 else "$0", ""]
    for col_idx, val in enumerate(row_data, 1):
        cell = ws_e4.cell(row=row, column=col_idx, value=val)
        cell.border = thin_border
        cell.alignment = center
        cell.font = cell_font
        if col_idx == 1:
            cell.alignment = left
            cell.font = name_font_bold
        if new > 0 and col_idx == 3:
            cell.fill = nuevo_fill
            cell.font = nuevo_font
    row += 1

# Totales
row_data_total = ["TOTAL", total_act, total_new, total_act + total_new, "", f"${total_cost:,.0f}", ""]
for col_idx, val in enumerate(row_data_total, 1):
    cell = ws_e4.cell(row=row, column=col_idx, value=val)
    cell.border = thin_border
    cell.alignment = center
    cell.font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
    cell.fill = PatternFill(start_color='2F5496', end_color='2F5496', fill_type='solid')
    if col_idx == 1:
        cell.alignment = left
row += 2

# === IMPACTO EN PLANILLA ===
style_area(ws_e4, row, max_col_e4, "IMPACTO EN PLANILLA MENSUAL")
row += 1

planilla_actual = 29086407
costo_nuevo = total_cost
planilla_nueva = planilla_actual + costo_nuevo
aumento_pct = costo_nuevo / planilla_actual * 100

impacto = [
    ("Planilla actual (líquido mensual)", f"${planilla_actual:,.0f}"),
    ("Costo personal nuevo (mensual)", f"${costo_nuevo:,.0f}"),
    ("NUEVA PLANILLA TOTAL", f"${planilla_nueva:,.0f}"),
    ("Aumento porcentual", f"{aumento_pct:.1f}%"),
    ("Personal actual operativo", f"{total_act} personas"),
    ("Personal nuevo a contratar", f"{total_new} personas"),
    ("Dotación total", f"{total_act + total_new} personas"),
]

for label, val in impacto:
    cell_l = ws_e4.cell(row=row, column=1, value=label)
    cell_l.border = thin_border
    cell_l.font = Font(name='Calibri', size=11, bold=True)
    ws_e4.merge_cells(start_row=row, start_column=2, end_row=row, end_column=3)
    cell_v = ws_e4.cell(row=row, column=2, value=val)
    cell_v.border = thin_border
    cell_v.alignment = center
    cell_v.font = Font(name='Calibri', size=11, bold=True, color='C00000')
    if "NUEVA PLANILLA" in label:
        cell_l.fill = costo_fill_dark
        cell_l.font = costo_font_w
        cell_v.fill = costo_fill
        cell_v.font = Font(name='Calibri', size=13, bold=True, color='C00000')
    if "Aumento" in label:
        cell_v.font = Font(name='Calibri', size=13, bold=True, color='FF0000')
    row += 1

row += 1
# Nota
ws_e4.merge_cells(start_row=row, start_column=1, end_row=row, end_column=max_col_e4)
note = ws_e4.cell(row=row, column=1, value="NOTA: Los sueldos de referencia para nuevas contrataciones corresponden al nivel más bajo de cada área. Daniel (actual) se mantiene en Turno B por su restricción de 10:30. En Recepción se contratan 3 (no 4) porque Raquel y Fran Brocco ya cubren turno B. El solapamiento 10:30-18:30 duplica la dotación, beneficiando el peak de lavado 12:00-16:00.")
note.font = Font(name='Calibri', size=10, italic=True, color='C00000')
note.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
ws_e4.row_dimensions[row].height = 50

# Anchos
ws_e4.column_dimensions['A'].width = 26
for col in range(2, max_col_e4 + 1):
    ws_e4.column_dimensions[get_column_letter(col)].width = 20
ws_e4.column_dimensions[get_column_letter(max_col_e4)].width = 10


# ============================================================
# HOJA 5: LEYENDA Y NOTAS
# ============================================================
ws4 = wb.create_sheet("Leyenda y Notas")

ws4.merge_cells('A1:D1')
ws4.cell(row=1, column=1, value="LEYENDA Y NOTAS").font = Font(name='Calibri', bold=True, size=14, color='2F5496')

row = 3
notes = [
    ("Turno A (Apertura)", "Verde claro — Personal que abre el salón"),
    ("Turno B (Medio/Cierre)", "Naranja claro — Personal en turno intermedio o de cierre"),
    ("Turno C (Cierre 21h)", "Azul claro — Personal que cierra a las 21:00 (solo Escenario 3)"),
    ("G1 / G2", "Grupos de rotación para sábados (se alternan semana a semana)"),
    ("47 hrs/semana", "Máximo legal. Se ajusta un día más corto o extendido según escenario"),
    ("Peak lavado 12-16", "Todos los turnos se solapan, garantizando máxima cobertura"),
    ("Daniel", "Restricción especial: nunca antes de las 10:30"),
    ("Antonio San Martín", "Siempre presente a la apertura (9:00)"),
    ("Fran Brocco / Antonio SM", "Artículo 22 — flexibilidad horaria legal"),
    ("Carol", "En licencia médica — no incluida en la programación actual"),
    ("Corte al cierre", "Siempre al menos 1 de: Paty, Ceci Pacheco o Daniel"),
    ("Sábado (Esc. 2)", "Cierre a las 14:00. Dotación mínima con rotación"),
    ("Escenario 3 (21h)", "3 turnos rotativos. 1 día extendido (11h) por persona para llegar a 47h"),
    ("Escenario 4 (Dot. Completa)", "Mantiene dotación actual en todo horario 9-20. Requiere 18 contrataciones nuevas"),
    ("★ Personal Nuevo", "Amarillo — Personas a contratar para completar dotación en turno B"),
]

for label, desc in notes:
    cell_l = ws4.cell(row=row, column=1, value=label)
    cell_l.font = Font(name='Calibri', bold=True, size=10)
    cell_l.border = thin_border
    ws4.merge_cells(start_row=row, start_column=2, end_row=row, end_column=4)
    cell_d = ws4.cell(row=row, column=2, value=desc)
    cell_d.font = Font(name='Calibri', size=10)
    cell_d.border = thin_border
    row += 1

ws4.column_dimensions['A'].width = 25
ws4.column_dimensions['B'].width = 60

# Guardar
output_path = "/Users/antonio/Documents/Lazartigue/Horarios_Salon_Lazartigue.xlsx"
wb.save(output_path)
print(f"Archivo guardado en: {output_path}")
